#!/usr/bin/env python3
"""Create a colored Excel review of the V2 review-only reconstruction."""

from __future__ import annotations

from scripts.common.terminal_output import console

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DATA_PATH = ROOT / "data/processed/v2/quantity_4h_reconstructed_review.csv"
FLAGS_PATH = ROOT / "data/processed/v2/quantity_4h_reconstruction_flags.csv"
OUTPUT_PATH = ROOT / "data/processed/v2/quantity_4h_reconstructed_review_colored.xlsx"

FEATURE_START_INDEX = 2
FEATURE_COUNT = 9
SHEET_NAME = "reconstructed_review"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INTERPOLATED_FILL = PatternFill("solid", fgColor="FFF2CC")
REMAINING_MISSING_FILL = PatternFill("solid", fgColor="F8CBAD")
INTERPOLATED_FONT = Font(color="7F6000")
REMAINING_MISSING_FONT = Font(color="9C0006")


def parse_value(value: str):
    """Keep identifiers/time as text, write numeric water-quality values as numbers."""
    if value == "" or value.lower() == "nan":
        return None
    try:
        return float(value)
    except ValueError:
        return value


def styled_cell(ws, value, fill=None, font=None, header: bool = False):
    cell = WriteOnlyCell(ws, value=value)
    if header:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        cell.alignment = Alignment(vertical="center")
    return cell


def build_colored_workbook() -> Path:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(SHEET_NAME)
    ws.freeze_panes = "A2"

    row_count = 0
    reconstructed_count = 0
    remaining_missing_count = 0

    with DATA_PATH.open("r", encoding="utf-8-sig", newline="") as data_file, FLAGS_PATH.open(
        "r", encoding="utf-8-sig", newline=""
    ) as flags_file:
        data_reader = csv.reader(data_file)
        flags_reader = csv.reader(flags_file)
        data_header = next(data_reader)
        flags_header = next(flags_reader)
        expected_status_headers = [f"{name}__status" for name in data_header[FEATURE_START_INDEX:]]
        if flags_header[FEATURE_START_INDEX:] != expected_status_headers:
            raise ValueError("Interpolation flag columns do not match data feature columns.")

        ws.append([styled_cell(ws, value, header=True) for value in data_header])
        row_count = 1

        for row_number, (data_row, flag_row) in enumerate(zip(data_reader, flags_reader), start=2):
            if data_row[0] != flag_row[0]:
                raise ValueError(f"Station mismatch at row {row_number}: {data_row[0]} vs {flag_row[0]}")
            out_row = []
            for idx, value in enumerate(data_row):
                parsed = parse_value(value) if idx >= FEATURE_START_INDEX else value
                status = ""
                if FEATURE_START_INDEX <= idx < FEATURE_START_INDEX + FEATURE_COUNT:
                    status = flag_row[idx]
                if status == "reconstructed":
                    out_row.append(styled_cell(ws, parsed, INTERPOLATED_FILL, INTERPOLATED_FONT))
                    reconstructed_count += 1
                elif status == "remaining_missing":
                    out_row.append(styled_cell(ws, parsed, REMAINING_MISSING_FILL, REMAINING_MISSING_FONT))
                    remaining_missing_count += 1
                else:
                    out_row.append(styled_cell(ws, parsed))
            ws.append(out_row)
            row_count += 1

        extra_data = next(data_reader, None)
        extra_flags = next(flags_reader, None)
        if extra_data is not None or extra_flags is not None:
            raise ValueError("Data and interpolation flag files have different row counts.")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(data_header))}{row_count}"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 19
    for col_idx in range(FEATURE_START_INDEX + 1, FEATURE_START_INDEX + FEATURE_COUNT + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    console.print(f"saved: {OUTPUT_PATH}")
    console.print(f"rows: {row_count - 1}")
    console.print(f"reconstructed colored cells: {reconstructed_count}")
    console.print(f"remaining missing colored cells: {remaining_missing_count}")
    return OUTPUT_PATH


def verify_workbook(path: Path) -> None:
    wb = load_workbook(path, read_only=True)
    ws = wb[SHEET_NAME]
    rows = ws.iter_rows()
    header = next(rows)
    row_count = 1 + sum(1 for _ in rows)
    console.print(f"verified sheets: {wb.sheetnames}")
    console.print(f"verified dimensions: {row_count} rows x {len(header)} columns")


def main() -> int:
    path = build_colored_workbook()
    verify_workbook(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
